from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import login_required, current_user
from models import db, Competitor
from utils.groq_ai import generate_competitors
from utils.pdf_generator import generate_competitor_pdf
import json

competitors_bp = Blueprint('competitors', __name__, url_prefix='/competitors')

@competitors_bp.route('/')
@login_required
def index():
    cid     = current_user.company_id
    page    = request.args.get('page', 1, type=int)
    search  = request.args.get('search', '')
    per_page = 20

    query = Competitor.query.filter_by(company_id=cid)
    if search:
        query = query.filter(Competitor.name.ilike(f'%{search}%'))
    
    pagination = query.order_by(Competitor.similarity_score.desc()).paginate(
        page=page, per_page=per_page, error_out=False)

    return render_template('competitors/index.html',
                           competitors=pagination.items,
                           pagination=pagination,
                           search=search)


@competitors_bp.route('/generate', methods=['POST'])
@login_required
def generate():
    company = current_user.company
    if not company.description:
        flash("Veuillez d'abord compléter la description de votre entreprise.", 'warning')
        return redirect(url_for('auth.profile'))

    keywords = []
    if company.keywords:
        try:
            keywords = json.loads(company.keywords)
        except Exception:
            keywords = []

    count = request.form.get('count', 20, type=int)
    count = min(count, 100)  # max 100 per call

    competitors_data, error = generate_competitors(
        company.name, company.description, company.country,
        company.sector, keywords, count
    )

    if error:
        flash(f"Erreur IA : {error}", 'error')
        return redirect(url_for('competitors.index'))

    added = 0
    skipped = 0
    for c in competitors_data:
        if not isinstance(c, dict):
            continue
        name = (c.get('name') or '').strip()
        if not name:
            continue
        # Deduplication: case-insensitive name check
        existing = Competitor.query.filter_by(company_id=company.id).filter(
            db.func.lower(Competitor.name) == name.lower()
        ).first()
        if existing:
            skipped += 1
            continue
        comp = Competitor(
            company_id     = company.id,
            name           = name,
            website        = c.get('website', ''),
            country        = c.get('country', company.country),
            city           = c.get('city', ''),
            sector         = c.get('sector', ''),
            activities     = c.get('activities', ''),
            products       = c.get('products', ''),
            services       = c.get('services', ''),
            employees_count= c.get('employees_count', ''),
            founded_year   = c.get('founded_year'),
            linkedin_url   = c.get('linkedin_url', ''),
            google_rating  = c.get('google_rating'),
            similarity_score = c.get('similarity_score', 50),
            revenue_estimate = c.get('revenue_estimate', ''),
            swot_analysis  = c.get('swot_analysis', ''),
            source         = 'ai'
        )
        db.session.add(comp)
        added += 1

    db.session.commit()
    flash(f'{added} concurrents générés et ajoutés avec succès !', 'success')
    return redirect(url_for('competitors.index'))


@competitors_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    if request.method == 'POST':
        comp = Competitor(
            company_id      = current_user.company_id,
            name            = request.form.get('name'),
            website         = request.form.get('website'),
            country         = request.form.get('country'),
            city            = request.form.get('city'),
            sector          = request.form.get('sector'),
            activities      = request.form.get('activities'),
            products        = request.form.get('products'),
            services        = request.form.get('services'),
            employees_count = request.form.get('employees_count'),
            founded_year    = request.form.get('founded_year') or None,
            linkedin_url    = request.form.get('linkedin_url'),
            facebook_url    = request.form.get('facebook_url'),
            google_rating   = request.form.get('google_rating') or None,
            similarity_score= request.form.get('similarity_score') or 50,
            revenue_estimate= request.form.get('revenue_estimate'),
            notes           = request.form.get('notes'),
            source          = 'manual'
        )
        db.session.add(comp)
        db.session.commit()
        flash('Concurrent ajouté avec succès.', 'success')
        return redirect(url_for('competitors.index'))
    return render_template('competitors/form.html', competitor=None)


@competitors_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    comp = Competitor.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    if request.method == 'POST':
        comp.name            = request.form.get('name', comp.name)
        comp.website         = request.form.get('website', comp.website)
        comp.country         = request.form.get('country', comp.country)
        comp.city            = request.form.get('city', comp.city)
        comp.sector          = request.form.get('sector', comp.sector)
        comp.activities      = request.form.get('activities', comp.activities)
        comp.products        = request.form.get('products', comp.products)
        comp.services        = request.form.get('services', comp.services)
        comp.employees_count = request.form.get('employees_count', comp.employees_count)
        comp.founded_year    = request.form.get('founded_year') or comp.founded_year
        comp.linkedin_url    = request.form.get('linkedin_url', comp.linkedin_url)
        comp.facebook_url    = request.form.get('facebook_url', comp.facebook_url)
        comp.google_rating   = request.form.get('google_rating') or comp.google_rating
        comp.similarity_score= request.form.get('similarity_score', comp.similarity_score)
        comp.revenue_estimate= request.form.get('revenue_estimate', comp.revenue_estimate)
        comp.notes           = request.form.get('notes', comp.notes)
        db.session.commit()
        flash('Concurrent mis à jour.', 'success')
        return redirect(url_for('competitors.index'))
    return render_template('competitors/form.html', competitor=comp)


@competitors_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete(id):
    comp = Competitor.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    db.session.delete(comp)
    db.session.commit()
    flash('Concurrent supprimé.', 'success')
    return redirect(url_for('competitors.index'))


@competitors_bp.route('/<int:id>/pdf')
@login_required
def pdf(id):
    comp = Competitor.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    buffer = generate_competitor_pdf(comp)
    filename = f"concurrent_{comp.name.replace(' ', '_')}.pdf"
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


@competitors_bp.route('/<int:id>/view')
@login_required
def view(id):
    comp = Competitor.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    swot = None
    if comp.swot_analysis:
        try:
            swot = json.loads(comp.swot_analysis)
        except Exception:
            swot = {'raw': comp.swot_analysis}
    return render_template('competitors/view.html', competitor=comp, swot=swot)


# ─── EXPORT XLSX ──────────────────────────────────────────────
@competitors_bp.route('/export/xlsx')
@login_required
def export_xlsx():
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    from models import Competitor

    cid = current_user.company_id
    items = Competitor.query.filter_by(company_id=cid)\
        .order_by(Competitor.similarity_score.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Concurrents"

    hdr_fill = PatternFill("solid", fgColor="1e429f")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    headers = ['Nom','Site Web','Pays','Ville','Secteur','Activités',
               'Produits','Services','Employés','Fondé','Score Similarité %',
               'CA Estimé','Note Google','LinkedIn','Source']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    for row, comp in enumerate(items, 2):
        ws.cell(row=row, column=1,  value=comp.name)
        ws.cell(row=row, column=2,  value=comp.website or '')
        ws.cell(row=row, column=3,  value=comp.country or '')
        ws.cell(row=row, column=4,  value=comp.city or '')
        ws.cell(row=row, column=5,  value=comp.sector or '')
        ws.cell(row=row, column=6,  value=comp.activities or '')
        ws.cell(row=row, column=7,  value=comp.products or '')
        ws.cell(row=row, column=8,  value=comp.services or '')
        ws.cell(row=row, column=9,  value=comp.employees_count or '')
        ws.cell(row=row, column=10, value=comp.founded_year or '')
        ws.cell(row=row, column=11, value=comp.similarity_score or '')
        ws.cell(row=row, column=12, value=comp.revenue_estimate or '')
        ws.cell(row=row, column=13, value=comp.google_rating or '')
        ws.cell(row=row, column=14, value=comp.linkedin_url or '')
        ws.cell(row=row, column=15, value='IA' if comp.source=='ai' else 'Manuel')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or '')) for c in col) + 4, 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name=f"concurrents_{__import__('datetime').datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── BULK DELETE ──────────────────────────────────────────────
@competitors_bp.route('/bulk-delete', methods=['POST'])
@login_required
def bulk_delete():
    from models import Competitor
    ids = request.form.getlist('selected_ids[]')
    if not ids:
        flash('Aucun concurrent sélectionné.', 'warning')
        return redirect(url_for('competitors.index'))
    count = 0
    for cid in ids:
        c = Competitor.query.filter_by(id=int(cid), company_id=current_user.company_id).first()
        if c:
            db.session.delete(c)
            count += 1
    db.session.commit()
    flash(f'{count} concurrent(s) supprimé(s).', 'success')
    return redirect(url_for('competitors.index'))