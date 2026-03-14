from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import login_required, current_user
from models import db, Prospect
from utils.groq_ai import generate_prospects
from utils.pdf_generator import generate_prospect_pdf
import json

prospects_bp = Blueprint('prospects', __name__, url_prefix='/prospects')

STATUS_LABELS = {
    'new': 'Nouveau', 'contacted': 'Contacté', 'replied': 'A répondu',
    'opportunity': 'Opportunité', 'interested': 'Intéressé',
    'quoted': 'Devis envoyé', 'won': 'Gagné', 'lost': 'Perdu',
    'delivered': 'Livré', 'maintenance': 'Maintenance'
}

@prospects_bp.route('/')
@login_required
def index():
    cid     = current_user.company_id
    page    = request.args.get('page', 1, type=int)
    search  = request.args.get('search', '')
    status  = request.args.get('status', '')
    sector  = request.args.get('sector', '')
    per_page = 20

    query = Prospect.query.filter_by(company_id=cid)
    if search:
        query = query.filter(Prospect.company_name.ilike(f'%{search}%'))
    if status:
        query = query.filter_by(status=status)
    if sector:
        query = query.filter(Prospect.sector.ilike(f'%{sector}%'))

    pagination = query.order_by(Prospect.relevance_score.desc()).paginate(
        page=page, per_page=per_page, error_out=False)

    # Get distinct sectors for filter
    sectors = db.session.query(Prospect.sector).filter_by(company_id=cid)\
        .distinct().filter(Prospect.sector != None).all()
    sectors = [s[0] for s in sectors if s[0]]

    return render_template('prospects/index.html',
                           prospects=pagination.items,
                           pagination=pagination,
                           search=search,
                           status_filter=status,
                           sector_filter=sector,
                           sectors=sectors,
                           status_labels=STATUS_LABELS)


@prospects_bp.route('/generate', methods=['POST'])
@login_required
def generate():
    company = current_user.company
    if not company.description:
        flash("Veuillez d'abord compléter la description de votre entreprise.", 'warning')
        return redirect(url_for('auth.profile'))

    icp_profile = {}
    if company.icp_profile:
        try:
            icp_profile = json.loads(company.icp_profile)
        except Exception:
            pass

    count = request.form.get('count', 20, type=int)
    count = min(count, 100)

    prospects_data, error = generate_prospects(
        company.name, company.description, company.country,
        company.sector, icp_profile, count
    )

    if error:
        flash(f"Erreur IA : {error}", 'error')
        return redirect(url_for('prospects.index'))

    added = 0
    skipped = 0
    for p in prospects_data:
        if not isinstance(p, dict):
            continue
        name = (p.get('company_name') or '').strip()
        if not name:
            continue
        # Deduplication: case-insensitive company name check
        existing = Prospect.query.filter_by(company_id=company.id).filter(
            db.func.lower(Prospect.company_name) == name.lower()
        ).first()
        if existing:
            skipped += 1
            continue
        prospect = Prospect(
            company_id      = company.id,
            company_name    = name,
            sector          = p.get('sector', ''),
            sub_sector      = p.get('sub_sector', ''),
            country         = p.get('country', company.country),
            city            = p.get('city', ''),
            size            = p.get('size', ''),
            employees_count = p.get('employees_count', ''),
            contact_name    = p.get('contact_name', ''),
            contact_title   = p.get('contact_title', ''),
            email           = p.get('email', ''),
            phone           = p.get('phone', ''),
            website         = p.get('website', ''),
            linkedin_url    = p.get('linkedin_url', ''),
            why_relevant    = p.get('why_relevant', ''),
            relevance_score = p.get('relevance_score', 50),
            source          = 'ai'
        )
        db.session.add(prospect)
        added += 1

    db.session.commit()
    msg = f'{added} nouveau(x) prospect(s) ajouté(s).'
    if skipped:
        msg += f' {skipped} doublon(s) ignoré(s).'
    flash(msg, 'success' if added > 0 else 'warning')
    return redirect(url_for('prospects.index'))


@prospects_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    if request.method == 'POST':
        prospect = Prospect(
            company_id      = current_user.company_id,
            company_name    = request.form.get('company_name'),
            sector          = request.form.get('sector'),
            sub_sector      = request.form.get('sub_sector'),
            country         = request.form.get('country'),
            city            = request.form.get('city'),
            size            = request.form.get('size'),
            employees_count = request.form.get('employees_count'),
            contact_name    = request.form.get('contact_name'),
            contact_title   = request.form.get('contact_title'),
            email           = request.form.get('email'),
            phone           = request.form.get('phone'),
            website         = request.form.get('website'),
            linkedin_url    = request.form.get('linkedin_url'),
            why_relevant    = request.form.get('why_relevant'),
            relevance_score = request.form.get('relevance_score', 50),
            notes           = request.form.get('notes'),
            source          = 'manual'
        )
        db.session.add(prospect)
        db.session.commit()
        flash('Prospect ajouté avec succès.', 'success')
        return redirect(url_for('prospects.index'))
    return render_template('prospects/form.html', prospect=None)


@prospects_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    p = Prospect.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    if request.method == 'POST':
        p.company_name    = request.form.get('company_name', p.company_name)
        p.sector          = request.form.get('sector', p.sector)
        p.sub_sector      = request.form.get('sub_sector', p.sub_sector)
        p.country         = request.form.get('country', p.country)
        p.city            = request.form.get('city', p.city)
        p.size            = request.form.get('size', p.size)
        p.employees_count = request.form.get('employees_count', p.employees_count)
        p.contact_name    = request.form.get('contact_name', p.contact_name)
        p.contact_title   = request.form.get('contact_title', p.contact_title)
        p.email           = request.form.get('email', p.email)
        p.phone           = request.form.get('phone', p.phone)
        p.website         = request.form.get('website', p.website)
        p.linkedin_url    = request.form.get('linkedin_url', p.linkedin_url)
        p.why_relevant    = request.form.get('why_relevant', p.why_relevant)
        p.relevance_score = request.form.get('relevance_score', p.relevance_score)
        p.notes           = request.form.get('notes', p.notes)
        p.status          = request.form.get('status', p.status)
        db.session.commit()
        flash('Prospect mis à jour.', 'success')
        return redirect(url_for('prospects.index'))
    return render_template('prospects/form.html', prospect=p, status_labels=STATUS_LABELS)


@prospects_bp.route('/<int:id>/status', methods=['POST'])
@login_required
def update_status(id):
    p = Prospect.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    new_status = request.form.get('status')
    if new_status in STATUS_LABELS:
        p.status = new_status
        db.session.commit()
        # If replied → create opportunity prompt
        if new_status == 'replied':
            flash(f'Statut mis à jour. {p.company_name} a répondu — créez une opportunité !', 'info')
        else:
            flash(f'Statut mis à jour : {STATUS_LABELS[new_status]}', 'success')
    return redirect(request.referrer or url_for('prospects.index'))


@prospects_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete(id):
    p = Prospect.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    db.session.delete(p)
    db.session.commit()
    flash('Prospect supprimé.', 'success')
    return redirect(url_for('prospects.index'))


@prospects_bp.route('/<int:id>/pdf')
@login_required
def pdf(id):
    p = Prospect.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    buffer = generate_prospect_pdf(p)
    filename = f"prospect_{p.company_name.replace(' ', '_')}.pdf"
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


@prospects_bp.route('/<int:id>/view')
@login_required
def view(id):
    from models import Opportunity, Quote, Order, EmailLog, EmailCampaign, MaintenanceOpportunity
    p = Prospect.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()

    # Complete history
    opportunities   = Opportunity.query.filter_by(prospect_id=p.id).order_by(Opportunity.created_at.desc()).all()
    email_logs      = EmailLog.query.filter_by(prospect_id=p.id).order_by(EmailLog.sent_at.desc()).all()
    maintenance_ops = MaintenanceOpportunity.query.filter_by(prospect_id=p.id).order_by(MaintenanceOpportunity.created_at.desc()).all()

    # Build unified timeline
    timeline = []
    for opp in opportunities:
        timeline.append({'date': opp.created_at, 'type': 'opportunity', 'obj': opp,
                         'label': f"Opportunité créée : {opp.title}",
                         'icon': 'fas fa-star', 'color': '#1a56db'})
        for q in opp.quotes:
            timeline.append({'date': q.created_at, 'type': 'quote', 'obj': q,
                              'label': f"Devis {q.quote_number} — {q.total_ttc:.0f} TND",
                              'icon': 'fas fa-file-alt', 'color': '#7c3aed'})
        for o in opp.orders:
            timeline.append({'date': o.ordered_at, 'type': 'order', 'obj': o,
                              'label': f"Bon de commande {o.order_number} — {o.amount:.0f} TND",
                              'icon': 'fas fa-shopping-cart', 'color': '#059669'})
    for el in email_logs:
        camp = EmailCampaign.query.get(el.campaign_id)
        timeline.append({'date': el.sent_at, 'type': 'email', 'obj': el,
                          'label': f"Email envoyé : {camp.subject if camp else '—'}",
                          'sub': '✅ Lu' if el.opened_at else ('💬 Répondu' if el.replied_at else '📤 Envoyé'),
                          'icon': 'fas fa-envelope', 'color': '#d97706'})
    for mo in maintenance_ops:
        timeline.append({'date': mo.created_at, 'type': 'maintenance', 'obj': mo,
                          'label': f"Maintenance : {mo.title}",
                          'icon': 'fas fa-tools', 'color': '#0891b2'})

    timeline.sort(key=lambda x: x['date'] or __import__("datetime").datetime.min, reverse=True)

    return render_template('prospects/view.html',
        prospect=p,
        status_labels=STATUS_LABELS,
        opportunities=opportunities,
        email_logs=email_logs,
        maintenance_ops=maintenance_ops,
        timeline=timeline,
    )


# ─── EXPORT XLSX ──────────────────────────────────────────────
@prospects_bp.route('/export/xlsx')
@login_required
def export_xlsx():
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file

    cid = current_user.company_id
    prospects = Prospect.query.filter_by(company_id=cid)\
        .order_by(Prospect.relevance_score.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospects"

    hdr_fill = PatternFill("solid", fgColor="1a56db")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)

    headers = ['Entreprise','Secteur','Sous-secteur','Pays','Ville','Taille',
               'Contact','Titre','Email','Téléphone','Site Web','LinkedIn',
               'Score','Statut','Source','Pertinence IA','Notes']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    status_labels = {
        'new':'Nouveau','contacted':'Contacté','replied':'A répondu',
        'opportunity':'Opportunité','interested':'Intéressé','quoted':'Devis',
        'won':'Gagné','lost':'Perdu','delivered':'Livré'
    }
    for row, p in enumerate(prospects, 2):
        ws.cell(row=row, column=1,  value=p.company_name)
        ws.cell(row=row, column=2,  value=p.sector or '')
        ws.cell(row=row, column=3,  value=p.sub_sector or '')
        ws.cell(row=row, column=4,  value=p.country or '')
        ws.cell(row=row, column=5,  value=p.city or '')
        ws.cell(row=row, column=6,  value=p.size or '')
        ws.cell(row=row, column=7,  value=p.contact_name or '')
        ws.cell(row=row, column=8,  value=p.contact_title or '')
        ws.cell(row=row, column=9,  value=p.email or '')
        ws.cell(row=row, column=10, value=p.phone or '')
        ws.cell(row=row, column=11, value=p.website or '')
        ws.cell(row=row, column=12, value=p.linkedin_url or '')
        ws.cell(row=row, column=13, value=p.relevance_score or '')
        ws.cell(row=row, column=14, value=status_labels.get(p.status, p.status))
        ws.cell(row=row, column=15, value='IA' if p.source=='ai' else 'Manuel')
        ws.cell(row=row, column=16, value=p.why_relevant or '')
        ws.cell(row=row, column=17, value=p.notes or '')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or '')) for c in col) + 4, 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name=f"prospects_{__import__('datetime').datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── BULK DELETE ──────────────────────────────────────────────
@prospects_bp.route('/bulk-delete', methods=['POST'])
@login_required
def bulk_delete():
    ids = request.form.getlist('selected_ids[]')
    if not ids:
        flash('Aucun prospect sélectionné.', 'warning')
        return redirect(url_for('prospects.index'))
    count = 0
    for pid in ids:
        p = Prospect.query.filter_by(id=int(pid), company_id=current_user.company_id).first()
        if p:
            db.session.delete(p)
            count += 1
    db.session.commit()
    flash(f'{count} prospect(s) supprimé(s).', 'success')
    return redirect(url_for('prospects.index'))