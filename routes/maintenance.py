from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from models import db, MaintenanceOpportunity, MaintenanceQuote, Prospect
from datetime import datetime
import uuid

maintenance_bp = Blueprint('maintenance', __name__, url_prefix='/maintenance')

PARK_TYPES = {
    'park_cvc':          '❄️ Parc CVC (Climatisation / Ventilation / Chauffage)',
    'park_froid':        '🧊 Parc Froid Commercial & Industriel',
    'park_materiel':     '⚙️ Parc Matériel Industriel',
    'park_informatique': '💻 Parc Informatique & Réseaux',
    'park_electrique':   '⚡ Parc Électrique & Énergie',
    'park_groupe':       '🔋 Groupes Électrogènes',
    'park_pompe':        '💧 Pompes & Hydraulique',
    'autre':             '📦 Autre / Multi-parcs',
}

STAGES = {
    'prospecting':  {'label': 'Prospection',   'icon': '🔍', 'color': '#64748b'},
    'quoted':       {'label': 'Devis envoyé',  'icon': '📄', 'color': '#1a56db'},
    'negotiating':  {'label': 'Négociation',   'icon': '🤝', 'color': '#d97706'},
    'won':          {'label': 'Contrat signé', 'icon': '✅', 'color': '#059669'},
    'active':       {'label': 'Contrat actif', 'icon': '🔧', 'color': '#0891b2'},
    'expired':      {'label': 'Expiré',        'icon': '⏰', 'color': '#dc2626'},
    'lost':         {'label': 'Perdu',         'icon': '❌', 'color': '#6b7280'},
}


# ── KANBAN INDEX ──────────────────────────────────────────────
@maintenance_bp.route('/')
@login_required
def index():
    cid = current_user.company_id
    pipeline = {}
    for stage, info in STAGES.items():
        opps = MaintenanceOpportunity.query.filter_by(company_id=cid, stage=stage)\
            .order_by(MaintenanceOpportunity.updated_at.desc()).all()
        pipeline[stage] = {'info': info, 'opps': opps}

    # KPIs
    total   = MaintenanceOpportunity.query.filter_by(company_id=cid).count()
    active  = MaintenanceOpportunity.query.filter_by(company_id=cid, stage='active').count()
    won     = MaintenanceOpportunity.query.filter_by(company_id=cid, stage='won').count()
    from sqlalchemy import func
    mrr = db.session.query(func.sum(MaintenanceOpportunity.value)).filter(
        MaintenanceOpportunity.company_id == cid,
        MaintenanceOpportunity.stage.in_(['won', 'active'])
    ).scalar() or 0

    return render_template('maintenance/index.html',
        pipeline=pipeline,
        park_types=PARK_TYPES,
        total=total, active=active, won=won, mrr=mrr
    )


# ── CREATE ────────────────────────────────────────────────────
@maintenance_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    cid = current_user.company_id
    prospects = Prospect.query.filter_by(company_id=cid)\
        .order_by(Prospect.company_name).all()

    if request.method == 'POST':
        nf  = request.form.get('next_followup')
        nv  = request.form.get('next_visit')
        opp = MaintenanceOpportunity(
            company_id       = cid,
            prospect_id      = request.form.get('prospect_id') or None,
            title            = request.form.get('title'),
            maint_type       = request.form.get('maint_type', 'preventive'),
            park_type        = request.form.get('park_type', 'autre'),
            park_description = request.form.get('park_description'),
            nb_equipments    = request.form.get('nb_equipments') or None,
            site_address     = request.form.get('site_address'),
            value            = request.form.get('value') or None,
            currency         = request.form.get('currency', 'TND'),
            contract_duration = request.form.get('contract_duration', 12),
            visits_per_year  = request.form.get('visits_per_year', 4),
            stage            = 'prospecting',
            probability      = request.form.get('probability', 40),
            notes            = request.form.get('notes'),
            assigned_to      = current_user.id,
            next_followup    = datetime.strptime(nf, '%Y-%m-%d') if nf else None,
            next_visit       = datetime.strptime(nv, '%Y-%m-%d') if nv else None,
        )
        db.session.add(opp)
        db.session.commit()
        flash('Opportunité de maintenance créée !', 'success')
        return redirect(url_for('maintenance.view', id=opp.id))

    return render_template('maintenance/new.html',
        prospects=prospects, park_types=PARK_TYPES)


# ── VIEW DETAIL ───────────────────────────────────────────────
@maintenance_bp.route('/<int:id>')
@login_required
def view(id):
    opp = MaintenanceOpportunity.query.filter_by(
        id=id, company_id=current_user.company_id).first_or_404()
    return render_template('maintenance/view.html',
        opp=opp, park_types=PARK_TYPES, stages=STAGES)


# ── EDIT ──────────────────────────────────────────────────────
@maintenance_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    opp = MaintenanceOpportunity.query.filter_by(
        id=id, company_id=current_user.company_id).first_or_404()
    prospects = Prospect.query.filter_by(company_id=current_user.company_id)\
        .order_by(Prospect.company_name).all()

    if request.method == 'POST':
        nf = request.form.get('next_followup')
        nv = request.form.get('next_visit')
        opp.title            = request.form.get('title', opp.title)
        opp.maint_type       = request.form.get('maint_type', opp.maint_type)
        opp.park_type        = request.form.get('park_type', opp.park_type)
        opp.park_description = request.form.get('park_description', opp.park_description)
        opp.nb_equipments    = request.form.get('nb_equipments') or opp.nb_equipments
        opp.site_address     = request.form.get('site_address', opp.site_address)
        opp.value            = request.form.get('value') or opp.value
        opp.currency         = request.form.get('currency', opp.currency)
        opp.contract_duration = request.form.get('contract_duration', opp.contract_duration)
        opp.visits_per_year  = request.form.get('visits_per_year', opp.visits_per_year)
        opp.probability      = request.form.get('probability', opp.probability)
        opp.stage            = request.form.get('stage', opp.stage)
        opp.notes            = request.form.get('notes', opp.notes)
        opp.prospect_id      = request.form.get('prospect_id') or opp.prospect_id
        if nf:
            opp.next_followup = datetime.strptime(nf, '%Y-%m-%d')
        if nv:
            opp.next_visit = datetime.strptime(nv, '%Y-%m-%d')
        if opp.stage == 'lost':
            opp.lost_reason = request.form.get('lost_reason', '')
        db.session.commit()
        flash('Opportunité mise à jour.', 'success')
        return redirect(url_for('maintenance.view', id=opp.id))

    return render_template('maintenance/edit.html',
        opp=opp, prospects=prospects, park_types=PARK_TYPES, stages=STAGES)


# ── QUICK STAGE ───────────────────────────────────────────────
@maintenance_bp.route('/<int:id>/stage', methods=['POST'])
@login_required
def update_stage(id):
    opp = MaintenanceOpportunity.query.filter_by(
        id=id, company_id=current_user.company_id).first_or_404()
    new_stage = request.form.get('stage')
    if new_stage in STAGES:
        opp.stage = new_stage
        if new_stage == 'lost':
            opp.lost_reason = request.form.get('lost_reason', '')
        db.session.commit()
        flash(f'Stade mis à jour : {STAGES[new_stage]["label"]}', 'success')
    return redirect(request.referrer or url_for('maintenance.index'))


# ── DELETE ────────────────────────────────────────────────────
@maintenance_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete(id):
    opp = MaintenanceOpportunity.query.filter_by(
        id=id, company_id=current_user.company_id).first_or_404()
    db.session.delete(opp)
    db.session.commit()
    flash('Opportunité de maintenance supprimée.', 'success')
    return redirect(url_for('maintenance.index'))


# ── ADD QUOTE ─────────────────────────────────────────────────
@maintenance_bp.route('/<int:id>/quote', methods=['POST'])
@login_required
def add_quote(id):
    opp = MaintenanceOpportunity.query.filter_by(
        id=id, company_id=current_user.company_id).first_or_404()

    total_ht  = float(request.form.get('total_ht', 0) or 0)
    tva_rate  = float(request.form.get('tva', 19) or 19)
    tva_amt   = total_ht * tva_rate / 100
    total_ttc = total_ht + tva_amt

    vu = request.form.get('valid_until')
    q  = MaintenanceQuote(
        opportunity_id = opp.id,
        quote_number   = f"MQ-{datetime.now().strftime('%Y%m')}-{uuid.uuid4().hex[:4].upper()}",
        total_ht       = total_ht,
        tva            = tva_rate,
        total_ttc      = total_ttc,
        status         = 'sent',
        notes          = request.form.get('notes'),
        valid_until    = datetime.strptime(vu, '%Y-%m-%d') if vu else None,
    )
    db.session.add(q)
    if opp.stage == 'prospecting':
        opp.stage = 'quoted'
    db.session.commit()
    flash(f'Devis {q.quote_number} créé.', 'success')
    return redirect(url_for('maintenance.view', id=opp.id))


# ── EXPORT XLSX ───────────────────────────────────────────────
@maintenance_bp.route('/export')
@login_required
def export_xlsx():
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file

    cid  = current_user.company_id
    opps = MaintenanceOpportunity.query.filter_by(company_id=cid)\
        .order_by(MaintenanceOpportunity.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance CRM"

    header_fill = PatternFill("solid", fgColor="0891b2")
    hdr_font    = Font(color="FFFFFF", bold=True, size=11)
    headers = ['Titre','Type','Parc','Client','Site','Nb Équip.','Valeur TND',
               'Durée (mois)','Visites/an','Stade','Probabilité %','Prochaine visite','Notes']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    for row, o in enumerate(opps, 2):
        ws.cell(row=row, column=1,  value=o.title)
        ws.cell(row=row, column=2,  value='Préventive' if o.maint_type=='preventive' else 'Curative' if o.maint_type=='curative' else 'Les deux')
        ws.cell(row=row, column=3,  value=PARK_TYPES.get(o.park_type, o.park_type or ''))
        ws.cell(row=row, column=4,  value=o.prospect.company_name if o.prospect else '')
        ws.cell(row=row, column=5,  value=o.site_address or '')
        ws.cell(row=row, column=6,  value=o.nb_equipments or '')
        ws.cell(row=row, column=7,  value=o.value or '')
        ws.cell(row=row, column=8,  value=o.contract_duration or 12)
        ws.cell(row=row, column=9,  value=o.visits_per_year or 4)
        ws.cell(row=row, column=10, value=STAGES.get(o.stage,{}).get('label',''))
        ws.cell(row=row, column=11, value=o.probability or '')
        ws.cell(row=row, column=12, value=o.next_visit.strftime('%d/%m/%Y') if o.next_visit else '')
        ws.cell(row=row, column=13, value=o.notes or '')

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name=f"maintenance_crm_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
